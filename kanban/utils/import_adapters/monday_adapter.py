"""
Monday.com Import Adapter
Handles import of Monday.com Excel exports (.xlsx).

Monday.com "Export board to Excel" produces .xlsx files with a specific structure:
- Row 1: Board name
- Row 2: Board description
- Row 3: Empty
- Row 4: Group name (e.g., "To-Do")
- Row 5: Column headers (Name, Owner, Status, Due date, ...)
- Row 6+: Task rows
- After tasks: optional date range summary row
- Pattern may repeat for additional groups
"""

import io
from typing import Dict, Any, Tuple, List, Optional
from collections import defaultdict

from .base_adapter import BaseImportAdapter, ImportResult, ImportError


class MondayAdapter(BaseImportAdapter):
    """
    Adapter for Monday.com Excel exports.
    """

    name = "Monday.com Adapter"
    supported_extensions = ['xlsx', 'xls']
    source_tool = "Monday.com"

    # Monday.com status-to-column mappings
    STATUS_COLUMN_MAP = {
        'working on it': 'Working on it',
        'done': 'Done',
        'stuck': 'Stuck',
        'not started': 'Not Started',
        '': 'Not Started',
    }

    # Monday.com default Kanban column order (matches their UI)
    STATUS_ORDER = {
        'Done': 0,
        'Working on it': 1,
        'Stuck': 2,
        'Not Started': 3,
    }

    # Map Monday.com header names to PrizmAI fields
    HEADER_MAP = {
        'name': 'title',
        'owner': 'assigned_to',
        'person': 'assigned_to',
        'status': 'column',
        'due date': 'due_date',
        'date': 'due_date',
        'timeline': 'timeline',
        'priority': 'priority',
        'notes': 'description',
        'text': 'description',
        'label': 'labels',
        'tags': 'labels',
        'numbers': 'progress',
    }

    def can_handle(self, data: Any, filename: str = None) -> Tuple[bool, float]:
        """
        Check if this is a Monday.com Excel export.
        """
        confidence = 0.0

        # Check filename extension
        if filename:
            ext = filename.lower().rsplit('.', 1)[-1] if '.' in filename else ''
            if ext in ('xlsx', 'xls'):
                confidence += 0.4

        # If we have binary data, try to read it as xlsx and look for Monday.com patterns
        if isinstance(data, bytes):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
                ws = wb.active

                rows = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    rows.append(row)
                    if i >= 10:
                        break
                wb.close()

                if len(rows) >= 5:
                    # Row 1 should have the board name (single cell value)
                    if rows[0] and rows[0][0] and isinstance(rows[0][0], str):
                        confidence += 0.1

                    # Row 5 (index 4) should have headers like Name, Owner, Status
                    header_row = rows[4] if len(rows) > 4 else None
                    if header_row:
                        headers_lower = [str(h).lower().strip() for h in header_row if h]
                        monday_headers = {'name', 'owner', 'status', 'due date', 'person', 'priority', 'date'}
                        matches = sum(1 for h in headers_lower if h in monday_headers)
                        if matches >= 2:
                            confidence += 0.4

            except Exception:
                pass

        return (confidence >= 0.5, min(confidence, 1.0))

    def parse(self, raw_data: Any) -> Dict[str, Any]:
        """Parse Monday.com Excel data into structured format."""
        if not isinstance(raw_data, bytes):
            raise ImportError("Monday.com adapter requires binary file data (.xlsx)")

        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for Excel imports. Run: pip install openpyxl")

        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw_data), read_only=True, data_only=True)
        except Exception as e:
            raise ImportError(f"Could not read Excel file: {e}")

        all_rows = []
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            all_rows.append(list(row))
        wb.close()

        if len(all_rows) < 5:
            raise ImportError("Excel file has too few rows to be a valid Monday.com export")

        # Extract board metadata from first rows
        board_name = str(all_rows[0][0]).strip() if all_rows[0][0] else 'Imported Board'
        board_description = str(all_rows[1][0]).strip() if len(all_rows) > 1 and all_rows[1][0] else ''

        # Parse groups and tasks
        # Monday.com exports can have multiple groups separated by group header rows
        groups = []
        current_group = None
        headers = None
        header_row_idx = None

        for idx, row in enumerate(all_rows):
            # Skip the first 3 rows (board name, description, empty)
            if idx < 3:
                continue

            # Clean row values
            values = [str(cell).strip() if cell is not None else '' for cell in row]
            non_empty = [v for v in values if v]

            # Skip completely empty rows
            if not non_empty:
                continue

            # Detect group header row: single non-empty cell in column A, no headers yet for this group,
            # OR it's a group separator
            if self._is_group_header(values, headers, header_row_idx, idx):
                current_group = values[0]
                headers = None
                header_row_idx = None
                continue

            # Detect header row: contains "Name" or multiple recognized headers
            if headers is None and self._is_header_row(values):
                headers = values
                header_row_idx = idx
                if current_group is None:
                    current_group = 'To Do'
                continue

            # Skip date range summary rows (e.g., "2026-03-18 to 2026-03-20")
            if headers and self._is_summary_row(values):
                continue

            # This should be a task row
            if headers and values[0]:
                task_row = {}
                for col_idx, header in enumerate(headers):
                    if header and col_idx < len(values):
                        task_row[header] = values[col_idx]
                task_row['_group'] = current_group or 'To Do'
                groups.append(task_row)

        return {
            'board_name': board_name,
            'board_description': board_description,
            'headers': headers or [],
            'tasks': groups,
        }

    def _is_group_header(self, values: List[str], headers, header_row_idx, current_idx) -> bool:
        """Detect if a row is a Monday.com group header (e.g., 'To-Do', 'In Progress')."""
        non_empty = [v for v in values if v]
        if len(non_empty) == 1 and values[0]:
            # Single value in first column - likely a group header
            # But not if it looks like a header row
            if not self._is_header_row(values):
                return True
        return False

    def _is_header_row(self, values: List[str]) -> bool:
        """Detect if a row is a column header row."""
        values_lower = [v.lower() for v in values if v]
        known_headers = {'name', 'owner', 'status', 'due date', 'person', 'priority',
                         'date', 'notes', 'text', 'timeline', 'label', 'tags', 'numbers',
                         'email', 'phone', 'link', 'subitems'}
        matches = sum(1 for v in values_lower if v in known_headers)
        return matches >= 2

    def _is_summary_row(self, values: List[str]) -> bool:
        """Detect date range summary rows like '2026-03-18 to 2026-03-20'."""
        for v in values:
            if v and ' to ' in v and any(char.isdigit() for char in v):
                return True
        return False

    def validate(self, parsed_data: Dict[str, Any]) -> bool:
        """Validate the parsed Monday.com data."""
        if not parsed_data.get('tasks'):
            raise ImportError("No tasks found in the Monday.com export")

        # Check that we have at least a name field
        headers = parsed_data.get('headers', [])
        headers_lower = [h.lower() for h in headers if h]
        if 'name' not in headers_lower:
            # Try to find any reasonable title-like column
            if headers:
                self.result.add_warning(f"No 'Name' column found. Using '{headers[0]}' as task name.")
            else:
                raise ImportError("Could not determine task name column from Monday.com export")

        return True

    def transform_to_prizmAI(self, parsed_data: Dict[str, Any]) -> ImportResult:
        """Transform Monday.com data to PrizmAI format."""
        result = ImportResult(
            source_format='xlsx',
            source_tool='Monday.com'
        )

        headers = parsed_data.get('headers', [])
        tasks_raw = parsed_data.get('tasks', [])

        # Build header mapping
        header_mapping = {}
        for header in headers:
            if not header:
                continue
            header_lower = header.lower().strip()
            if header_lower in self.HEADER_MAP:
                header_mapping[header] = self.HEADER_MAP[header_lower]

        # If no 'title' mapping found, use first column
        mapped_fields = set(header_mapping.values())
        if 'title' not in mapped_fields and headers:
            header_mapping[headers[0]] = 'title'

        result.field_mappings = {v: k for k, v in header_mapping.items()}

        # Collect columns and labels
        columns_seen = {}
        labels_seen = {}
        tasks = []

        for row_idx, row_data in enumerate(tasks_raw):
            task = self._transform_task(row_data, header_mapping, row_idx, columns_seen, labels_seen)
            if task:
                tasks.append(task)
                result.update_stats('tasks_imported')

        # Build column data
        for col_name, col_info in sorted(columns_seen.items(), key=lambda x: x[1]['order']):
            result.columns_data.append({
                'name': col_name,
                'position': col_info['order'],
                'temp_id': col_info['temp_id'],
            })
            result.update_stats('columns_imported')

        # Build label data
        for label_name, label_color in labels_seen.items():
            result.labels_data.append({
                'name': label_name,
                'color': label_color,
            })
            result.update_stats('labels_imported')

        result.tasks_data = tasks

        # Board data
        result.board_data = {
            'name': parsed_data.get('board_name', 'Imported from Monday.com'),
            'description': parsed_data.get('board_description', ''),
        }

        return result

    def _transform_task(self, row_data: Dict, header_mapping: Dict, row_idx: int,
                        columns_seen: Dict, labels_seen: Dict) -> Optional[Dict[str, Any]]:
        """Transform a single Monday.com row to a PrizmAI task."""

        def get_field(field_name: str) -> str:
            """Get value for a PrizmAI field from the row."""
            for header, mapped_field in header_mapping.items():
                if mapped_field == field_name and header in row_data:
                    return row_data[header]
            return ''

        # Get title
        title = get_field('title')
        if not title:
            return None

        # Determine column from Status field or group
        status = get_field('column')
        if status:
            # Normalize Monday.com statuses
            column_name = self.STATUS_COLUMN_MAP.get(status.lower().strip(), status.strip())
        else:
            # Fall back to the group name
            column_name = row_data.get('_group', 'To Do')

        if column_name not in columns_seen:
            # Use Monday.com's standard column order if known, otherwise append at end
            order = self.STATUS_ORDER.get(column_name, 100 + len(columns_seen))
            columns_seen[column_name] = {
                'order': order,
                'temp_id': f'col_{len(columns_seen)}',
            }

        # Process labels
        label_names = []
        labels_value = get_field('labels')
        if labels_value:
            for label in str(labels_value).split(','):
                label = label.strip()
                if label:
                    label_names.append(label)
                    if label not in labels_seen:
                        labels_seen[label] = self._generate_label_color(len(labels_seen))

        # Map priority
        priority = self._normalize_priority(get_field('priority'))

        # Parse due date
        due_date = self._parse_date(get_field('due_date'))

        # Parse timeline for start_date
        start_date = None
        timeline = get_field('timeline')
        if timeline and ' - ' in str(timeline):
            parts = str(timeline).split(' - ')
            start_date = self._parse_date(parts[0].strip())
            if not due_date:
                due_date = self._parse_date(parts[1].strip())

        # Build task data
        task_data = {
            'title': self._sanitize_string(title, max_length=200),
            'description': get_field('description') or '',
            'column_temp_id': columns_seen[column_name]['temp_id'],
            'position': row_idx,
            'priority': priority,
            'progress': self._status_to_progress(status),
            'due_date': due_date,
            'start_date': start_date,
            'assigned_to_username': get_field('assigned_to') or None,
            'label_names': label_names,
            'external_id': None,
            'complexity_score': 5,
        }

        return task_data

    def _status_to_progress(self, status: str) -> int:
        """Convert Monday.com status to progress percentage."""
        if not status:
            return 0
        status_lower = status.lower().strip()
        progress_map = {
            'done': 100,
            'working on it': 50,
            'stuck': 25,
            'not started': 0,
        }
        return progress_map.get(status_lower, 0)

    def _generate_label_color(self, index: int) -> str:
        """Generate a color for a label based on index."""
        colors = [
            '#61bd4f', '#f2d600', '#ff9f1a', '#eb5a46', '#c377e0',
            '#0079bf', '#00c2e0', '#51e898', '#ff78cb', '#344563',
        ]
        return colors[index % len(colors)]
